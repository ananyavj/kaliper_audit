#core/plan_loader.py
import json
import re
from pathlib import Path
from typing import Any, List, Dict
import openpyxl
from core.schemas import ConditionalRule, TrackingEventSpec

_SKIP_SHEETS = {
    "overview", "global props", "data dictionary",
    "summary & kpis", "pql scoring model", "engagement benchmarks",
    "simulator coverage gaps", "simulator_coverage_gaps",
}


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _split_properties_list(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "—"}:
        return []
    parts = re.split(r"[,/\n]", text)
    return [p.strip() for p in parts if p.strip()]


def _parse_conditional_rules(raw: list[dict]) -> list[ConditionalRule]:
    """
    D1: parse the optional "conditional_rules" array from a JSON event spec.

    Expected shape per rule:
      {
        "when_property": "payment_method",
        "when_value": "emi",
        "then_property": "emi_bank",
        "then_allowed_values": ["HDFC", "ICICI", "SBI"]   // optional
      }
    """
    rules = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        when_prop = item.get("when_property", "").strip()
        when_val  = item.get("when_value")
        then_prop = item.get("then_property", "").strip()
        if not when_prop or when_val is None or not then_prop:
            continue   # skip malformed rules silently
        rules.append(ConditionalRule(
            when_property=when_prop,
            when_value=when_val,
            then_property=then_prop,
            then_allowed_values=item.get("then_allowed_values", []),
        ))
    return rules


def _load_from_dict(data: dict) -> list[TrackingEventSpec]:
    specs = []
    for event in data.get("events", []):
        # D1: read conditional_rules if present; default to empty list
        conditional_rules = _parse_conditional_rules(
            event.get("conditional_rules", [])
        )
        specs.append(TrackingEventSpec(
            name=event["name"],
            required_properties=event.get("required_properties", []),
            property_types=event.get("property_types", {}),
            identity_required=event.get("identity_required", False),
            allowed_previous_events=event.get("allowed_previous_events", []),
            allowed_values=event.get("allowed_values", {}),
            conditional_rules=conditional_rules,
        ))
    return specs


def _load_from_json(path: Path) -> list[TrackingEventSpec]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return _load_from_dict(data)


def _load_from_excel(path: Path) -> list[TrackingEventSpec]:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_specs = []
    seen_names = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in _SKIP_SHEETS:
            continue

        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row
        header_idx = None
        for idx, r in enumerate(rows[:10]):
            cells = [str(c).strip().lower() if c is not None else "" for c in r]
            if any("event name" in cell for cell in cells):
                header_idx = idx
                headers = cells
                break
        if header_idx is None:
            continue

        col_map = {h: i for i, h in enumerate(headers)}

        for r in rows[header_idx + 1:]:
            if not r or len(r) <= col_map.get("event name", 0):
                continue
            name_cell = r[col_map["event name"]]
            if name_cell is None:
                continue
            name_str = str(name_cell).strip()
            if not name_str or name_str.startswith("▸") or name_str.startswith("▶"):
                continue

            # Key properties
            key_props_idx = col_map.get("key properties")
            key_props_raw = r[key_props_idx] if key_props_idx is not None and key_props_idx < len(r) else ""
            required_properties = _split_properties_list(key_props_raw)

            # Property types
            prop_type_idx = col_map.get("property type")
            prop_types_raw = r[prop_type_idx] if prop_type_idx is not None and prop_type_idx < len(r) else ""
            types_list = _split_properties_list(prop_types_raw)

            _type_map = {
                "str": "string", "string": "string",
                "int": "number", "integer": "number", "number": "number", "float": "number",
                "bool": "boolean", "boolean": "boolean",
                "iso8601": "string", "iso 8601": "string",
                "object": "object", "dict": "object",
                "array": "array", "list": "array",
            }
            property_types = {}
            for i, prop in enumerate(required_properties):
                t = types_list[i] if i < len(types_list) else "string"
                property_types[prop] = _type_map.get(t.strip().lower(), "string")

            # Also capture any additional typed properties listed in a separate
            # "optional properties" or "all properties" column if present.
            # Previously only required_properties got types; optional-but-typed
            # properties (e.g. revenue: number) were silently dropped, so
            # wrong_property_type checks never fired for them.
            all_props_idx = col_map.get("optional properties") or col_map.get("all properties")
            if all_props_idx is not None and all_props_idx < len(r):
                all_props_raw = r[all_props_idx]
                all_props_list = _split_properties_list(all_props_raw)
                # Pair with types_list continuing after required_properties entries
                offset = len(required_properties)
                for j, prop in enumerate(all_props_list):
                    if prop and prop not in property_types:
                        t_idx = offset + j
                        t = types_list[t_idx] if t_idx < len(types_list) else "string"
                        property_types[prop] = _type_map.get(t.strip().lower(), "string")

            # Identity required
            id_fields_idx = col_map.get("identity fields")
            identity_fields = []
            if id_fields_idx is not None and id_fields_idx < len(r):
                identity_fields = _split_properties_list(r[id_fields_idx])
            identity_required = len(identity_fields) > 0 or "user_id" in required_properties

            # Allowed previous events
            prev_events_idx = col_map.get("allowed previous events")
            allowed_previous_events = []
            if prev_events_idx is not None and prev_events_idx < len(r):
                allowed_previous_events = _split_properties_list(r[prev_events_idx])

            # Allowed values
            # Note: no default currency injection here (A1 fix — see plan_normalizer.py)
            allowed_values: dict = {}

            # D1: Excel plans don't have a conditional_rules column yet — empty list
            conditional_rules: list[ConditionalRule] = []

            spec = TrackingEventSpec(
                name=name_str,
                required_properties=required_properties,
                property_types=property_types,
                identity_required=identity_required,
                allowed_previous_events=allowed_previous_events,
                allowed_values=allowed_values,
                conditional_rules=conditional_rules,
            )
            if spec.name not in seen_names:
                all_specs.append(spec)
                seen_names.add(spec.name)

    return all_specs


def load_tracking_plan(path: str) -> list[TrackingEventSpec]:
    plan_path = Path(path)
    suffix = plan_path.suffix.lower()
    if suffix == ".json":
        return _load_from_json(plan_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_from_excel(plan_path)
    raise ValueError("Tracking plan must be a .json, .xlsx, or .xlsm file.")


def load_tracking_plan_from_dict(data: dict) -> list[TrackingEventSpec]:
    return _load_from_dict(data)
